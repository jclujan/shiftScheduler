# -*- coding: utf-8 -*-
"""
exportar.py
===========
Genera los archivos descargables (Excel con colores y CSV plano) a partir de
la programacion calculada por motor.py.

Convencion de colores en Excel:
  - Amarillo: el operador esta negociado (no esta en su turno o libranza fija).
  - Rojo: el turno no alcanzo el minimo de operadores (regla incumplida).
"""

import io
import calendar
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from motor import TURNOS, NOMBRES_DIA, MESES_ABREV, MINIMO

# Colores reutilizables
AMARILLO = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ROJO = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
GRIS = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
BORDE = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def _tabla_larga(programacion):
    """
    Construye una tabla 'larga' (una fila por dia y turno). Sirve tanto para
    el CSV como para una de las hojas de Excel.
    """
    filas = []
    for fecha in sorted(programacion):
        for turno in TURNOS:
            datos = programacion[fecha][turno]
            negociados_txt = "; ".join(
                f"{op} ({tipo})" for op, tipo in datos["negociados"].items()
            )
            filas.append({
                "Fecha": fecha.strftime("%Y-%m-%d"),
                "Dia": NOMBRES_DIA[fecha.weekday()],
                "Turno": turno,
                "Cantidad": len(datos["operadores"]),
                "Operadores": ", ".join(datos["operadores"]),
                "Negociaciones": negociados_txt,
                "Cumple_minimo": "Si" if datos["cumple"] else "No",
            })
    return pd.DataFrame(filas)


def exportar_csv(programacion):
    """Devuelve los bytes de un CSV plano (sin colores)."""
    df = _tabla_larga(programacion)
    return df.to_csv(index=False).encode("utf-8-sig")


def exportar_excel(programacion, anio, mes, operadores=None,
                   vacaciones=None, rank_turno=None, rank_libranza=None):
    """
    Devuelve los bytes de un Excel con estas hojas:
      - 'Programacion': tabla larga con colores por celda.
      - 'Calendario': vista de calendario del periodo con los tres turnos por dia.
      - 'Configuracion': los datos de entrada (solo si se pasan 'operadores').
    """
    wb = Workbook()

    # ------------------------------------------------------------------ #
    # Hoja 1: tabla larga con colores
    # ------------------------------------------------------------------ #
    hoja = wb.active
    hoja.title = "Programacion"
    encabezados = ["Fecha", "Dia", "Turno", "Cantidad",
                   "Operadores", "Negociaciones", "Cumple_minimo"]
    hoja.append(encabezados)
    for celda in hoja[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        celda.alignment = Alignment(horizontal="center")

    for fecha in sorted(programacion):
        for turno in TURNOS:
            datos = programacion[fecha][turno]
            negociados_txt = "; ".join(
                f"{op} ({tipo})" for op, tipo in datos["negociados"].items()
            )
            hoja.append([
                fecha.strftime("%Y-%m-%d"),
                NOMBRES_DIA[fecha.weekday()],
                turno,
                len(datos["operadores"]),
                ", ".join(datos["operadores"]),
                negociados_txt,
                "Si" if datos["cumple"] else "No",
            ])
            fila = hoja.max_row
            # Rojo si el turno no cumple el minimo.
            if not datos["cumple"]:
                for col in range(1, 8):
                    hoja.cell(row=fila, column=col).fill = ROJO
            # Amarillo en la celda de negociaciones si hubo alguna.
            elif negociados_txt:
                hoja.cell(row=fila, column=6).fill = AMARILLO

    anchos = {"A": 12, "B": 11, "C": 9, "D": 10, "E": 45, "F": 40, "G": 14}
    for col, ancho in anchos.items():
        hoja.column_dimensions[col].width = ancho

    # ------------------------------------------------------------------ #
    # Hoja 2: vista de calendario mensual
    # ------------------------------------------------------------------ #
    cal = wb.create_sheet("Calendario")
    cal.append([f"Programacion de turnos - {calendar.month_name[mes]} {anio}"])
    cal.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    cal["A1"].font = Font(bold=True, size=14)
    cal["A1"].alignment = Alignment(horizontal="center")

    # Encabezado de dias de la semana (empezando en lunes).
    cal.append(NOMBRES_DIA)
    for celda in cal[2]:
        celda.font = Font(bold=True)
        celda.fill = GRIS
        celda.alignment = Alignment(horizontal="center")

    # matriz de semanas del periodo (fechas reales, ya vienen completas lun-dom)
    fechas = sorted(programacion)
    semanas = [fechas[i:i + 7] for i in range(0, len(fechas), 7)]

    fila_actual = 3
    for semana in semanas:
        # Cada dia ocupa una celda con: numero + los 3 turnos y su gente.
        contenidos = []
        estilos = []  # (tiene_rojo, tiene_amarillo)
        for fecha in semana:
            dia = programacion[fecha]
            # Etiqueta del dia; si es de un mes vecino se marca con el mes.
            if fecha.month != mes:
                etiqueta = f"{fecha.day} {MESES_ABREV[fecha.month - 1]}"
            else:
                etiqueta = str(fecha.day)
            lineas = [etiqueta]
            tiene_rojo = tiene_amarillo = False
            for turno in TURNOS:
                datos = dia[turno]
                marca = ""
                if not datos["cumple"]:
                    marca = " [ROJO]"
                    tiene_rojo = True
                elif datos["negociados"]:
                    marca = " [*]"
                    tiene_amarillo = True
                lineas.append(f"{turno[:3]}: {', '.join(datos['operadores'])}{marca}")
            contenidos.append("\n".join(lineas))
            estilos.append((tiene_rojo, tiene_amarillo))

        cal.append(contenidos)
        for col, (tiene_rojo, tiene_amarillo) in enumerate(estilos, start=1):
            celda = cal.cell(row=fila_actual, column=col)
            celda.alignment = Alignment(wrap_text=True, vertical="top")
            celda.border = BORDE
            if tiene_rojo:
                celda.fill = ROJO
            elif tiene_amarillo:
                celda.fill = AMARILLO
        cal.row_dimensions[fila_actual].height = 90
        fila_actual += 1

    for col in "ABCDEFG":
        cal.column_dimensions[col].width = 26

    # Leyenda al pie
    fila_actual += 1
    cal.cell(row=fila_actual, column=1, value="[*] = requiere negociacion (amarillo)")
    cal.cell(row=fila_actual, column=1).fill = AMARILLO
    cal.cell(row=fila_actual + 1, column=1, value="[ROJO] = no se cumplio el minimo de 3")
    cal.cell(row=fila_actual + 1, column=1).fill = ROJO

    # ------------------------------------------------------------------ #
    # Hoja 3 (opcional): Configuracion, o sea los datos de entrada usados.
    # Se incluye si se pasan operadores, para que el Excel lleve tambien los
    # inputs (turno y libranza de cada quien, rankings y vacaciones).
    # ------------------------------------------------------------------ #
    if operadores is not None:
        vacaciones = vacaciones or {}
        rank_turno = rank_turno or []
        rank_libranza = rank_libranza or []
        pos_turno = {op: i + 1 for i, op in enumerate(rank_turno)}
        pos_libranza = {op: i + 1 for i, op in enumerate(rank_libranza)}

        conf = wb.create_sheet("Configuracion")
        cabecera = ["Anio", "Mes", "Operador", "Turno", "Libranza",
                    "RankTurno", "RankLibranza", "Vacaciones"]
        conf.append(cabecera)
        for celda in conf[1]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
            celda.alignment = Alignment(horizontal="center")

        for op, info in operadores.items():
            dias = sorted(vacaciones.get(op, set()))
            vac_txt = "; ".join(d.isoformat() for d in dias)
            conf.append([
                anio, mes, op, info["turno"], info["libranza"],
                pos_turno.get(op, ""), pos_libranza.get(op, ""), vac_txt,
            ])

        anchos_conf = {"A": 8, "B": 6, "C": 18, "D": 10, "E": 16,
                       "F": 11, "G": 13, "H": 40}
        for col, ancho in anchos_conf.items():
            conf.column_dimensions[col].width = ancho

    # ------------------------------------------------------------------ #
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
